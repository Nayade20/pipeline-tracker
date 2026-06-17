import requests
import smtplib
import os
from datetime import datetime, timedelta, timezone
import pytz
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

# ============================================================
# CONFIGURACIÓN — editar estos valores
# ============================================================
HUBSPOT_TOKEN      = "pat-na1-cd6a882d-248a-45d8-99fb-6d1e71f77845"
EMAIL_REMITENTE    = "info@zentralcom.com"
EMAIL_CONTRASENA   = "ptF3dCC2TBQPmV"
EMAIL_DESTINATARIO = ["nayade.barrutieta@zentralcom.com"]
DIAS_ATRAS         = 7
# ============================================================

HEADERS = {
    "Authorization": f"Bearer {HUBSPOT_TOKEN}",
    "Content-Type": "application/json"
}

def _fill(color): return PatternFill("solid", fgColor=color)
def _font(color="000000", bold=False, size=9, underline=None):
    return Font(name="Arial", size=size, bold=bold, color=color, underline=underline)
def _border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left"): return Alignment(horizontal=h, vertical="center")

HDR_FILLS = {
    "dark":   _fill("1F3864"), "blue":   _fill("2E75B6"),
    "green":  _fill("375623"), "orange": _fill("C55A11"),
    "gray":   _fill("595959"), "purple": _fill("7030A0"),
    "teal":   _fill("1F6B75"),
}
ROW_FILLS = {
    "ganado":  _fill("E2EFDA"), "perdido": _fill("FCE4D6"),
    "cliente": _fill("DDEEFF"), "alt":     _fill("F5F5F5"),
    "white":   _fill("FFFFFF"),
}
ACT_COLOR = {"Reunión":"375623","Email":"1F4E79","Llamada":"833C00","Nota":"7030A0","—":"888888"}
ACT_BG    = {"Reunión":"EAF3DE","Email":"E6F1FB","Llamada":"FAEEDA","Nota":"EEEDFE","—":"F1EFE8"}
TIPO_MAP  = {"MEETING":"Reunión","EMAIL":"Email","CALL":"Llamada","NOTE":"Nota","TASK":"Tarea"}


def cargar_owners():
    owners = {
        "74753477":  "Inigo Mangas Insausti",
        "75326442":  "Jose Ramon Mendibil",
        "75326441":  "Alejandro Alzas Moreno",
        "102123176": "Iban Ibanez",
        "75887314":  "Onura Onura",
        "75631012":  "Rafael Quintanilla",
        "83853436":  "David Rivero Gomes",
        "91245262":  "Olatz Alkorta",
        "1186290725":"Nayade Barrutieta",
    }
    try:
        after = None
        while True:
            params = {"limit": 100}
            if after:
                params["after"] = after
            r = requests.get("https://api.hubapi.com/crm/v3/owners", headers=HEADERS, params=params)
            if r.status_code != 200:
                break
            data = r.json()
            for o in data.get("results", []):
                oid = str(o.get("id") or o.get("ownerId") or "")
                if not oid:
                    continue
                name = f"{o.get('firstName','').strip()} {o.get('lastName','').strip()}".strip()
                if not name:
                    name = o.get("email", "")
                if oid and name:
                    owners[oid] = name
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after:
                break
    except Exception as e:
        print(f"Advertencia: no se pudieron cargar owners desde API: {e}")
    return owners


def cargar_etapas_pipelines():
    r = requests.get("https://api.hubapi.com/crm/v3/pipelines/deals", headers=HEADERS)
    stage_map, pipeline_map = {}, {}
    if r.status_code == 200:
        for p in r.json().get("results", []):
            pipeline_map[p["id"]] = p["label"]
            for s in p["stages"]:
                stage_map[s["id"]] = s["label"]
    return stage_map, pipeline_map


def obtener_ultima_actividad(deal_id):
    url = f"https://api.hubapi.com/engagements/v1/engagements/associated/deal/{deal_id}/paged?limit=100"
    r = requests.get(url, headers=HEADERS)
    ultima, tipo = None, "—"
    if r.status_code != 200:
        return ultima, tipo
    ahora = datetime.now()
    for item in r.json().get("results", []):
        eng = item.get("engagement", {})
        t   = eng.get("type", "")
        ts  = eng.get("timestamp")
        if t in TIPO_MAP and ts:
            fecha = datetime.fromtimestamp(ts / 1000)
            if fecha <= ahora and (not ultima or fecha > ultima):
                ultima, tipo = fecha, TIPO_MAP[t]
    return ultima, tipo


def obtener_deals_semana(dias=7):
    desde_ts = int((datetime.now() - timedelta(days=dias)).timestamp() * 1000)
    url = "https://api.hubapi.com/crm/v3/objects/deals/search"
    payload = {
        "filterGroups": [{"filters": [{
            "propertyName": "notes_last_updated",
            "operator": "GTE",
            "value": str(desde_ts)
        }]}],
        "properties": ["dealname","dealstage","pipeline","hubspot_owner_id",
                       "notes_last_updated","hs_object_id","amount"],
        "limit": 100,
        "sorts": [{"propertyName": "notes_last_updated", "direction": "DESCENDING"}]
    }
    deals, after = [], None
    while True:
        if after:
            payload["after"] = after
        r = requests.post(url, headers=HEADERS, json=payload)
        if r.status_code != 200:
            print(f"❌ Error al obtener deals: {r.status_code} {r.text}")
            break
        data = r.json()
        deals.extend(data.get("results", []))
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break
    return deals


def _hdr_cell(ws, row, col, value, fill_key):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = HDR_FILLS[fill_key]
    c.alignment = _align("center")
    c.border    = _border()


def generar_excel(deals_data, fecha_str):
    wb = Workbook()

    # ── Hoja 1: Negocios activos ─────────────────────────────
    ws = wb.active
    ws.title = "Negocios Activos"

    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = f"Negocios con actividad — semana hasta {fecha_str}  ({len(deals_data)} negocios)"
    tc.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    tc.fill      = HDR_FILLS["dark"]
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    hdrs = ["ID Negocio","Nombre del negocio","Comercial","Pipeline","Etapa",
            "Tipo actividad","Últ. actividad HubSpot","Enlace HubSpot"]
    hfks = ["dark","blue","blue","green","orange","purple","teal","gray"]
    ws.row_dimensions[2].height = 22
    for ci, (h, fk) in enumerate(zip(hdrs, hfks), 1):
        _hdr_cell(ws, 2, ci, h, fk)

    for ri, d in enumerate(deals_data, 3):
        pip   = d["pipeline"]
        etapa = d["etapa"].lower()

        if "ganado" in etapa:    rf = ROW_FILLS["ganado"]
        elif "perdido" in etapa: rf = ROW_FILLS["perdido"]
        elif pip == "Clientes":  rf = ROW_FILLS["cliente"]
        elif ri % 2 == 0:        rf = ROW_FILLS["white"]
        else:                    rf = ROW_FILLS["alt"]

        url = f"https://app.hubspot.com/contacts/48253960/record/0-3/{d['id']}"
        row = [d["id"], d["nombre"], d["propietario"], pip, d["etapa"],
               d["tipo_act"], d["fecha_mod"], "Ver en HubSpot"]

        ws.row_dimensions[ri].height = 16
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = rf
            cell.border    = _border()
            cell.alignment = _align()
            if ci == 4:
                cell.font = _font("1F4E79" if val=="Clientes" else "7B3F00", bold=True)
            elif ci == 6:
                cell.font = _font(ACT_COLOR.get(val,"888888"), bold=True)
                cell.fill = _fill(ACT_BG.get(val,"F1EFE8"))
            elif ci == 8:
                cell.hyperlink = url
                cell.font = _font("0563C1", underline="single")
            else:
                cell.font = _font()

    for i, w in enumerate([16,46,22,12,36,16,22,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{len(deals_data)+2}"

    # ── Hoja 2: Resumen por comercial ────────────────────────
    ws2 = wb.create_sheet("Resumen por comercial")
    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value = f"Resumen por comercial — semana hasta {fecha_str}"
    t2.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t2.fill = HDR_FILLS["blue"]
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 22

    h2  = ["Comercial","Negocios","Clientes","Lead","Reuniones","Llamadas","Emails","Notas"]
    h2f = ["blue","dark","green","orange","teal","teal","teal","teal"]
    ws2.row_dimensions[2].height = 20
    for ci, (h, fk) in enumerate(zip(h2, h2f), 1):
        _hdr_cell(ws2, 2, ci, h, fk)

    od = {}
    for d in deals_data:
        o = d["propietario"]
        if o not in od:
            od[o] = {"t":0,"c":0,"l":0,"r":0,"ll":0,"e":0,"n":0}
        od[o]["t"] += 1
        if d["pipeline"] == "Clientes": od[o]["c"] += 1
        else:                           od[o]["l"] += 1
        at = d["tipo_act"]
        if at == "Reunión":   od[o]["r"]  += 1
        elif at == "Llamada": od[o]["ll"] += 1
        elif at == "Email":   od[o]["e"]  += 1
        elif at == "Nota":    od[o]["n"]  += 1

    for ri, (owner, d) in enumerate(sorted(od.items(), key=lambda x: -x[1]["t"]), 3):
        rf2 = ROW_FILLS["white"] if ri % 2 == 0 else ROW_FILLS["alt"]
        ws2.row_dimensions[ri].height = 16
        for ci, val in enumerate([owner,d["t"],d["c"],d["l"],d["r"],d["ll"],d["e"],d["n"]], 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = _font(); c.fill = rf2; c.border = _border()
            c.alignment = _align() if ci == 1 else _align("center")

    for i, w in enumerate([24,12,12,12,12,12,12,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 3: Resumen por etapa ────────────────────────────
    ws3 = wb.create_sheet("Resumen por etapa")
    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = f"Resumen por etapa — semana hasta {fecha_str}"
    t3.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t3.fill = HDR_FILLS["orange"]
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 22
    for ci, (h, fk) in enumerate(zip(["Etapa","Pipeline","Negocios"],["orange","green","dark"]), 1):
        _hdr_cell(ws3, 2, ci, h, fk)
    ws3.row_dimensions[2].height = 20
    sc = Counter((d["etapa"], d["pipeline"]) for d in deals_data)
    for ri, ((etapa, pip), cnt) in enumerate(sorted(sc.items(), key=lambda x: -x[1]), 3):
        rf3 = ROW_FILLS["white"] if ri % 2 == 0 else ROW_FILLS["alt"]
        ws3.row_dimensions[ri].height = 16
        for ci, val in enumerate([etapa, pip, cnt], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = _font(); c.fill = rf3; c.border = _border()
            c.alignment = _align() if ci == 1 else _align("center")
    for i, w in enumerate([38,14,12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 4: Resumen por tipo de actividad ────────────────
    ws4 = wb.create_sheet("Resumen por actividad")
    ws4.merge_cells("A1:B1")
    t4 = ws4["A1"]
    t4.value = "Resumen por tipo de actividad"
    t4.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t4.fill = HDR_FILLS["purple"]
    t4.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[1].height = 22
    for ci, (h, fk) in enumerate(zip(["Tipo actividad","Negocios"],["purple","dark"]), 1):
        _hdr_cell(ws4, 2, ci, h, fk)
    ws4.row_dimensions[2].height = 20
    ac = Counter(d["tipo_act"] for d in deals_data)
    for ri, (at, cnt) in enumerate(sorted(ac.items(), key=lambda x: -x[1]), 3):
        rf4 = ROW_FILLS["white"] if ri % 2 == 0 else ROW_FILLS["alt"]
        ws4.row_dimensions[ri].height = 16
        for ci, val in enumerate([at, cnt], 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.fill = rf4; c.border = _border()
            c.alignment = _align() if ci == 1 else _align("center")
            c.font = _font(ACT_COLOR.get(val,"000000"), bold=True) if ci == 1 else _font()
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 12

    archivo = f"hubspot_negocios_ultima_actividad_ultima_semana_{fecha_str}.xlsx"
    wb.save(archivo)
    return archivo


def enviar_email(archivo, fecha_str, n_deals):
    msg = EmailMessage()
    msg["Subject"] = f"📊 HubSpot Negocios Última Actividad Última Semana — {fecha_str} ({n_deals} negocios)"
    msg["From"]    = EMAIL_REMITENTE
    msg["To"]      = ", ".join(EMAIL_DESTINATARIO)
    msg.set_content(
        f"Hola,\n\n"
        f"Adjunto el informe semanal de negocios con actividad en los últimos {DIAS_ATRAS} días.\n\n"
        f"• Total negocios: {n_deals}\n"
        f"• Período: últimos {DIAS_ATRAS} días hasta {fecha_str}\n\n"
        f"Un saludo,\nZentralcom"
    )
    with open(archivo, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(archivo)
        )
    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_REMITENTE, EMAIL_CONTRASENA)
        smtp.send_message(msg)
    print(f"📧 Correo enviado a {EMAIL_DESTINATARIO}")


def main():
    fecha_str = datetime.now().strftime("%Y-%m-%d")
    print(f"🚀 Iniciando informe semanal HubSpot — {fecha_str}")

    stage_map, pipeline_map = cargar_etapas_pipelines()
    owner_map = cargar_owners()

    print(f"📥 Obteniendo negocios con actividad en los últimos {DIAS_ATRAS} días...")
    raw_deals = obtener_deals_semana(DIAS_ATRAS)
    print(f"   → {len(raw_deals)} negocios encontrados")

    deals_data = []
    for i, deal in enumerate(raw_deals, 1):
        props   = deal.get("properties", {})
        deal_id = props.get("hs_object_id", deal.get("id",""))
        sid     = props.get("dealstage", "")
        pip_id  = props.get("pipeline", "")
        oid     = str(props.get("hubspot_owner_id") or "")

        pip_label = pipeline_map.get(pip_id, pip_id)
        if "cliente" in pip_label.lower():
            pip_label = "Clientes"
        elif "lead" in pip_label.lower():
            pip_label = "Lead"

        # Última actividad con fecha y hora en zona horaria local (Europe/Madrid)
        fecha_mod_raw = props.get("notes_last_updated", "")
        if fecha_mod_raw:
            try:
                tz_madrid = pytz.timezone("Europe/Madrid")
                dt_utc = datetime.fromisoformat(fecha_mod_raw.replace("Z", "+00:00"))
                dt_local = dt_utc.astimezone(tz_madrid)
                offset = dt_local.strftime("%z")
                offset_fmt = f"GMT{offset[:3]}:{offset[3:]}" if len(offset) == 5 else f"GMT{offset}"
                fecha_mod = dt_local.strftime(f"%Y-%m-%d %H:%M {offset_fmt}")
            except Exception:
                fecha_mod = fecha_mod_raw[:16].replace("T", " ")
        else:
            fecha_mod = "—"

        print(f"   [{i}/{len(raw_deals)}] {props.get('dealname','?')[:40]}", end=" ", flush=True)
        ultima, tipo = obtener_ultima_actividad(deal_id)
        print(f"→ {tipo}")

        deals_data.append({
            "id":          deal_id,
            "nombre":      props.get("dealname",""),
            "propietario": owner_map.get(oid, oid),
            "pipeline":    pip_label,
            "etapa":       stage_map.get(sid, sid),
            "stage_id":    sid,
            "tipo_act":    tipo,
            "fecha_mod":   fecha_mod,
        })

    print(f"\n📊 Generando Excel...")
    archivo = generar_excel(deals_data, fecha_str)
    print(f"✅ Excel guardado: {archivo}")

    print(f"📧 Enviando correo...")
    enviar_email(archivo, fecha_str, n_deals=len(deals_data))
    print("✅ Proceso completado.")


if __name__ == "__main__":
    main()