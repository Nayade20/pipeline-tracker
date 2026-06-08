import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import time
import io
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─────────────────────────────────────────────
# CONFIGURACIÓN — pon aquí tu API Key
# ─────────────────────────────────────────────

# Lee la API Key y contraseña desde Streamlit Secrets (cloud) o directamente (local)
import streamlit as _st
try:
    HUBSPOT_API_KEY = _st.secrets["HUBSPOT_API_KEY"]
    APP_PASSWORD    = _st.secrets["APP_PASSWORD"]
except Exception:
    HUBSPOT_API_KEY = "pat-eu1-xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"  # ← solo para uso local
    APP_PASSWORD    = ""  # sin contraseña en local

OUTPUT_DIR = r"C:\Users\Administracion1\Zentralcom\Zentralcom S.L - Documentos\Administracion\NAYADE\CODIGOS PYTHON\informe automatico hubspot"

BASE_URL = "https://api.hubapi.com"

OWNER_NAMES = {
    "74753477":   "Iñigo Mangas",
    "75326441":   "Usuario 75326441",
    "75326442":   "José R. Mendibil",
    "75631012":   "Rafa Quintanilla",
    "75887314":   "Onura Onura",
    "83853436":   "David Rivero",
    "91245262":   "Olatz Alkorta",
    "92982282":   "Administración Zentralcom",
    "102123176":  "Iban Ibañez",
    "1186290725": "Nayade Barrutieta",
}

PIPELINE_NAMES = {
    "default":   "CLIENTES",
    "664092197": "LEAD",
}

STAGE_LABELS = {
    "960918614":  "Objetivo clientes familias nuevas",
    "960918615":  "Presentadas nuevas familias — venta cruzada",
    "960918616":  "Facturas recibidas (cliente)",
    "960918617":  "Petición más info (cliente)",
    "960918618":  "Solicitud presupuesto proveedor (cliente)",
    "960918619":  "Oferta recibida proveedor (cliente)",
    "960918620":  "Informe preparado (cliente)",
    "960918621":  "Informe presentado (cliente)",
    "990804832":  "Negocio ganado",
    "1077494083": "Seguimiento subcontratación",
    "996161568":  "Negocio perdido",
    "974964831":  "Nuevo",
    "974964832":  "Empresa cualificada",
    "974964833":  "Contacto clave identificado",
    "974964834":  "Contactado sin feedback",
    "991276872":  "Contactado sin interés",
    "974964835":  "Interesado — sin reunión",
    "1077424678": "Seguimiento subcontratación (lead)",
    "1185458868": "Seguimiento Onura",
    "1185458869": "Seguimiento Rafa",
    "974964836":  "Reunión realizada",
    "974964837":  "Facturas recibidas (lead)",
    "975190032":  "Petición más info (lead)",
    "975190033":  "Solicitud presupuesto proveedor (lead)",
    "975190034":  "Oferta recibida proveedor (lead)",
    "990996102":  "Informe preparado (lead)",
    "990996103":  "Informe presentado (lead)",
    "991037466":  "Cierre ganado",
    "991037467":  "Cierre perdido",
}

WON_STAGES  = {"990804832", "991037466"}
LOST_STAGES = {"996161568", "991037467"}
STALE_DAYS  = [14, 30, 60]

# ─────────────────────────────────────────────
# CLIENTE API HUBSPOT
# ─────────────────────────────────────────────

def hs_get(api_key, url, params=None):
    headers = {"Authorization": f"Bearer {api_key}"}
    resp = requests.get(url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()

def hs_post(api_key, url, payload):
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    resp = requests.post(url, headers=headers, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()

def test_connection(api_key):
    try:
        hs_get(api_key, f"{BASE_URL}/crm/v3/owners")
        return True
    except Exception:
        return False

def get_all_deals(api_key):
    props = [
        "dealname", "dealstage", "pipeline", "hubspot_owner_id",
        "amount", "createdate", "closedate",
        "hs_lastmodifieddate", "hs_last_activity_date", "notes_last_updated",
    ]
    all_deals = []
    params = {"limit": 100, "properties": ",".join(props)}
    while True:
        data = hs_get(api_key, f"{BASE_URL}/crm/v3/objects/deals", params)
        all_deals.extend(data.get("results", []))
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break
        params["after"] = after
        time.sleep(0.1)
    return all_deals

def get_real_activity(api_key, since_ts_ms, until_ts_ms):
    """
    Obtiene actividad real: llamadas, emails, reuniones, notas y tareas
    asociadas a negocios en el período indicado.
    Devuelve: {deal_id: {"owner_id": str, "types": set}}
    """
    activity_endpoints = {
        "calls":    "/crm/v3/objects/calls",
        "emails":   "/crm/v3/objects/emails",
        "meetings": "/crm/v3/objects/meetings",
        "notes":    "/crm/v3/objects/notes",
        # tasks excluido — no se considera actividad real
    }
    type_props = {
        "calls":    ["hs_timestamp", "hubspot_owner_id"],
        "emails":   ["hs_timestamp", "hubspot_owner_id"],
        "meetings": ["hs_timestamp", "hubspot_owner_id"],
        "notes":    ["hs_timestamp", "hubspot_owner_id"],
    }
    deal_activity = {}

    for act_type, url in activity_endpoints.items():
        # 1. Recoger actividades en el rango de fechas
        in_range = []
        after = None
        while True:
            params = {"limit": 100, "properties": ",".join(type_props[act_type])}
            if after:
                params["after"] = after
            try:
                data = hs_get(api_key, url, params)
            except Exception:
                break
            for item in data.get("results", []):
                ts_raw = item.get("properties", {}).get("hs_timestamp")
                if not ts_raw:
                    continue
                try:
                    ts_ms = int(ts_raw)
                except (ValueError, TypeError):
                    try:
                        ts_ms = int(pd.to_datetime(ts_raw, utc=True).timestamp() * 1000)
                    except Exception:
                        continue
                if since_ts_ms <= ts_ms <= until_ts_ms:
                    in_range.append(item)
            paging = data.get("paging", {})
            after = paging.get("next", {}).get("after")
            if not after:
                break
            time.sleep(0.1)

        if not in_range:
            continue

        # 2. Para cada actividad, buscar los negocios asociados
        owner_map = {
            item["id"]: item.get("properties", {}).get("hubspot_owner_id", "")
            for item in in_range
        }
        ids = list(owner_map.keys())

        for i in range(0, len(ids), 100):
            batch = ids[i:i + 100]
            assoc_url = f"{BASE_URL}/crm/v4/associations/{act_type}/deals/batch/read"
            payload = {"inputs": [{"id": aid} for aid in batch]}
            try:
                assoc_data = hs_post(api_key, assoc_url, payload)
                for result in assoc_data.get("results", []):
                    act_id   = str(result.get("from", {}).get("id", ""))
                    owner_id = owner_map.get(act_id, "")
                    for assoc in result.get("to", []):
                        deal_id = str(assoc.get("toObjectId") or assoc.get("id", ""))
                        if deal_id not in deal_activity:
                            deal_activity[deal_id] = {"owner_id": owner_id, "types": set()}
                        deal_activity[deal_id]["types"].add(act_type)
            except Exception:
                pass
            time.sleep(0.15)

    return deal_activity

def get_stage_history_bulk(api_key, deal_ids):
    results = {}
    for i in range(0, len(deal_ids), 10):
        batch = deal_ids[i:i + 10]
        payload = {
            "inputs": [{"id": str(did)} for did in batch],
            "propertiesWithHistory": ["dealstage"],
            "properties": ["dealname", "hubspot_owner_id", "pipeline"],
        }
        try:
            data = hs_post(api_key, f"{BASE_URL}/crm/v3/objects/deals/batch/read", payload)
            for item in data.get("results", []):
                did = item["id"]
                results[did] = {
                    "dealname":     item["properties"].get("dealname", ""),
                    "owner_id":     item["properties"].get("hubspot_owner_id", ""),
                    "pipeline":     item["properties"].get("pipeline", ""),
                    "stage_history": item.get("propertiesWithHistory", {}).get("dealstage", []),
                }
        except Exception:
            pass
        time.sleep(0.2)
    return results

def get_engagement_types(api_key, deal_ids):
    """
    Obtiene el tipo de última actividad por negocio usando el endpoint de engagements.
    Igual que el código de referencia: /engagements/v1/engagements/associated/deal/{id}
    """
    TIPO_MAP = {"MEETING":"Reunión","EMAIL":"Email","CALL":"Llamada","NOTE":"Nota"}
    # TASK excluido — las tareas no se consideran actividad real
    results = {}
    for deal_id in deal_ids:
        url = f"https://api.hubapi.com/engagements/v1/engagements/associated/deal/{deal_id}/paged?limit=100"
        try:
            data = hs_get(api_key, url)
            ultima, tipo = None, "—"
            ahora = datetime.now()
            for item in data.get("results", []):
                eng = item.get("engagement", {})
                t   = eng.get("type", "")
                ts  = eng.get("timestamp")
                if t in TIPO_MAP and ts:
                    fecha = datetime.fromtimestamp(ts / 1000)
                    if fecha <= ahora and (not ultima or fecha > ultima):
                        ultima, tipo = fecha, TIPO_MAP[t]
            results[str(deal_id)] = tipo
        except Exception:
            results[str(deal_id)] = "—"
        time.sleep(0.1)
    return results


# ─────────────────────────────────────────────
# ANÁLISIS DE DATOS
# ─────────────────────────────────────────────

def parse_dt(val):
    if not val:
        return pd.NaT
    try:
        return pd.to_datetime(val, utc=True).tz_convert(None)
    except Exception:
        return pd.NaT

def parse_deals(raw_deals):
    rows = []
    for d in raw_deals:
        p = d.get("properties", {})
        owner_id    = str(p.get("hubspot_owner_id") or "")
        pipeline_id = str(p.get("pipeline") or "")
        rows.append({
            "deal_id":            d["id"],
            "dealname":           p.get("dealname", ""),
            "dealstage":          p.get("dealstage", ""),
            "pipeline_id":        pipeline_id,
            "pipeline":           PIPELINE_NAMES.get(pipeline_id, pipeline_id),
            "owner_id":           owner_id,
            "owner":              OWNER_NAMES.get(owner_id, f"Owner {owner_id}"),
            "amount":             float(p.get("amount") or 0),
            "createdate":         parse_dt(p.get("createdate")),
            "closedate":          parse_dt(p.get("closedate")),
            "last_modified":      parse_dt(p.get("hs_lastmodifieddate")),
            "last_activity_date": parse_dt(p.get("hs_last_activity_date")),
            "notes_last_updated": parse_dt(p.get("notes_last_updated")),
        })
    return pd.DataFrame(rows)

def activity_by_owner(df, week_start, week_end):
    """
    Usa notes_last_updated — misma propiedad que usa HubSpot para 'Última actividad'.
    La tasa de actividad se calcula sobre negocios ABIERTOS (excluye ganados/perdidos).
    """
    active_mask = (
        df["notes_last_updated"].notna() &
        (df["notes_last_updated"] >= week_start) &
        (df["notes_last_updated"] <= week_end)
    )
    active    = df[active_mask]
    # df ya solo contiene negocios abiertos
    total_open = df.groupby("owner").size().reset_index(name="open_deals")
    act        = active.groupby("owner").size().reset_index(name="active_this_week")

    result = total_open.merge(act, on="owner", how="left").fillna(0)
    result["open_deals"]       = result["open_deals"].astype(int)
    result["active_this_week"] = result["active_this_week"].astype(int)
    result["total_deals"]      = result["open_deals"]  # compatibilidad con resto del código
    result["sin_actividad"]    = result["open_deals"] - result["active_this_week"]
    result["sin_actividad"]    = result["sin_actividad"].clip(lower=0)
    result["activity_rate"]    = (
        result["active_this_week"] / result["open_deals"] * 100
    ).where(result["open_deals"] > 0, 0).round(1)
    return result.sort_values("active_this_week", ascending=False)

def activity_detail_by_owner(df, real_activity):
    """Desglose de tipos de actividad por comercial (solo si se cargó actividad real)."""
    rows = []
    for deal_id, info in real_activity.items():
        owner_id = str(info.get("owner_id", ""))
        owner    = OWNER_NAMES.get(owner_id, f"Owner {owner_id}")
        deal_row = df[df["deal_id"] == deal_id]
        if not deal_row.empty:
            owner = deal_row.iloc[0]["owner"]
        for act_type in info.get("types", set()):
            rows.append({"owner": owner, "activity_type": act_type})
    if not rows:
        return pd.DataFrame(columns=["owner", "activity_type"])
    return pd.DataFrame(rows)

def stage_activity_summary(df, week_start, week_end):
    """Actividad por etapa usando notes_last_updated."""
    active = df[
        df["notes_last_updated"].notna() &
        (df["notes_last_updated"] >= week_start) &
        (df["notes_last_updated"] <= week_end)
    ]
    total  = df.groupby("dealstage").size().reset_index(name="total_deals")
    act    = active.groupby("dealstage").size().reset_index(name="active_this_week")
    result = total.merge(act, on="dealstage", how="left").fillna(0)
    result["active_this_week"] = result["active_this_week"].astype(int)
    result["stage_label"]      = result["dealstage"].map(STAGE_LABELS).fillna(result["dealstage"])
    return result.sort_values("active_this_week", ascending=False)

def stage_transitions(history_data, week_start, week_end):
    rows = []
    for deal_id, info in history_data.items():
        hist     = info.get("stage_history", [])
        owner    = OWNER_NAMES.get(str(info.get("owner_id", "")), f"Owner {info.get('owner_id','')}")
        pipeline = PIPELINE_NAMES.get(str(info.get("pipeline", "")), info.get("pipeline", ""))
        for i, entry in enumerate(hist):
            ts = parse_dt(entry.get("timestamp"))
            if ts and week_start <= ts <= week_end and i + 1 < len(hist):
                from_stage_id = hist[i+1].get("value","")
                to_stage_id   = entry.get("value","")

                # Ignorar si no hay etapa anterior (negocio recién creado)
                if not from_stage_id or not from_stage_id.strip():
                    continue

                # Ignorar si etapa anterior y nueva son iguales
                if from_stage_id == to_stage_id:
                    continue

                from_label = STAGE_LABELS.get(from_stage_id, from_stage_id)
                to_label   = STAGE_LABELS.get(to_stage_id, to_stage_id)

                rows.append({
                    "deal_id":    deal_id,
                    "dealname":   info.get("dealname", ""),
                    "owner":      owner,
                    "pipeline":   pipeline,
                    "from_stage": from_label,
                    "to_stage":   to_label,
                    "changed_at": ts,
                })
    cols = ["deal_id","dealname","owner","pipeline","from_stage","to_stage","changed_at"]
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)

def new_deals_by_stage(df, week_start, week_end):
    new    = df[(df["createdate"] >= week_start) & (df["createdate"] <= week_end)]
    result = new.groupby("dealstage").size().reset_index(name="new_deals")
    result["stage_label"] = result["dealstage"].map(STAGE_LABELS).fillna(result["dealstage"])
    return result.sort_values("new_deals", ascending=False)

def owner_stage_matrix(df):
    df2 = df.copy()
    df2["stage_label"] = df2["dealstage"].map(STAGE_LABELS).fillna(df2["dealstage"])
    pivot = df2.pivot_table(index="owner", columns="stage_label",
                            values="deal_id", aggfunc="count", fill_value=0)
    pivot["TOTAL"] = pivot.sum(axis=1)
    return pivot.reset_index()

# Etapas cerradas — excluidas del dashboard principal
CLOSED_STAGES = WON_STAGES | LOST_STAGES | {
    "991037466",  # Cierre ganado
    "991037467",  # Cierre perdido
    "990804832",  # Negocio ganado
    "996161568",  # Negocio perdido
    "991276872",  # Contactado sin interés (ambos pipelines)
}

def stale_deals(df, thresholds, reference_date):
    """Negocios abiertos (sin etapas de cierre) sin actividad desde hace X días."""
    open_df = df.copy()  # df_all ya solo contiene negocios abiertos
    open_df["ref_date"]      = open_df["notes_last_updated"].fillna(open_df["last_modified"])
    open_df["days_inactive"] = (reference_date - open_df["ref_date"]).dt.days
    results = {}
    for days in thresholds:
        stale = open_df[open_df["days_inactive"] >= days].copy()
        stale["stage_label"] = stale["dealstage"].map(STAGE_LABELS).fillna(stale["dealstage"])
        results[days] = stale.sort_values("days_inactive", ascending=False)
    return results

# ─────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1A3A5C")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
SUBHEAD_FILL = PatternFill("solid", start_color="2E75B6")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ACCENT_FILL  = PatternFill("solid", start_color="DEEAF1")
ALT_FILL     = PatternFill("solid", start_color="F5F9FD")
NORMAL_FONT  = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", size=10, bold=True)
GREEN_FILL   = PatternFill("solid", start_color="E2EFDA")
RED_FILL     = PatternFill("solid", start_color="FCE4D6")
ORANGE_FILL  = PatternFill("solid", start_color="FFF2CC")
THIN         = Side(style="thin", color="CCCCCC")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def xl_header(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1A3A5C")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, color="595959")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

def xl_style_headers(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32

def xl_style_data(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row+1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        for c in range(1, ncols+1):
            cell = ws.cell(r, c)
            cell.font   = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if fill.fill_type:
                cell.fill = fill

def xl_auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len+2, max_w))

def xl_sheet_resumen(wb, week_label, act_df, stage_df, trans_df, new_df, stale_data):
    ws = wb.create_sheet("📊 Resumen semanal")
    xl_header(ws, "Resumen semanal del pipeline", f"Semana: {week_label}", 4)
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, "INDICADORES CLAVE").fill = SUBHEAD_FILL
    ws.cell(r, 1).font = SUBHEAD_FONT
    r += 1
    total   = int(act_df["total_deals"].sum())
    active  = int(act_df["active_this_week"].sum())
    rate    = round(active/total*100, 1) if total else 0
    stale14 = len(stale_data.get(14, pd.DataFrame()))
    for label, val in [
        ("Total negocios",              total),
        ("Con actividad esta semana",   active),
        ("Tasa de actividad",           f"{rate}%"),
        ("Cambios de etapa detectados", len(trans_df)),
        ("Negocios nuevos",             int(new_df["new_deals"].sum()) if not new_df.empty else 0),
        ("Negocios estancados >14 días",stale14),
    ]:
        ws.cell(r, 1, label).font = NORMAL_FONT
        ws.cell(r, 2, val).font   = BOLD_FONT
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, "ACTIVIDAD POR COMERCIAL").fill = SUBHEAD_FILL
    ws.cell(r, 1).font = SUBHEAD_FONT
    r += 1
    for ci, h in enumerate(["Comercial","Total negocios","Con actividad","Tasa %"], 1):
        ws.cell(r, ci, h)
    xl_style_headers(ws, r, 4)
    ds = r+1
    for _, row in act_df.iterrows():
        r += 1
        ws.cell(r, 1, row["owner"])
        ws.cell(r, 2, int(row["total_deals"]))
        ws.cell(r, 3, int(row["active_this_week"]))
        ws.cell(r, 4, f"{row['activity_rate']}%")
        fill = GREEN_FILL if row["activity_rate"] >= 20 else (ORANGE_FILL if row["activity_rate"] >= 10 else RED_FILL)
        ws.cell(r, 3).fill = fill
    xl_style_data(ws, ds, r, 4)
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_actividad(wb, week_label, act_df):
    ws = wb.create_sheet("👤 Actividad comercial")
    xl_header(ws, "Actividad por comercial", f"Semana: {week_label}", 5)
    r = 4
    for ci, h in enumerate(["Comercial","Total negocios","Con actividad","Sin actividad","Tasa %"], 1):
        ws.cell(r, ci, h)
    xl_style_headers(ws, r, 5)
    ds = r+1
    for _, row in act_df.iterrows():
        r += 1
        ws.cell(r, 1, row["owner"])
        ws.cell(r, 2, int(row["total_deals"]))
        ws.cell(r, 3, int(row["active_this_week"]))
        ws.cell(r, 4, int(row["sin_actividad"]))
        ws.cell(r, 5, row["activity_rate"])
        fill = GREEN_FILL if row["activity_rate"] >= 20 else (ORANGE_FILL if row["activity_rate"] >= 10 else RED_FILL)
        ws.cell(r, 3).fill = fill
    xl_style_data(ws, ds, r, 5)
    chart = BarChart()
    chart.title = "Actividad esta semana por comercial"
    chart.height, chart.width = 12, 20
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=ds-1, max_row=r)
    cats     = Reference(ws, min_col=1, min_row=ds, max_row=r)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{r+3}")
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_etapas(wb, week_label, stage_df):
    ws = wb.create_sheet("📋 Etapas — actividad")
    xl_header(ws, "Actividad por etapa", f"Semana: {week_label}", 3)
    r = 4
    for ci, h in enumerate(["Etapa","Total negocios","Con actividad esta semana"], 1):
        ws.cell(r, ci, h)
    xl_style_headers(ws, r, 3)
    ds = r+1
    for _, row in stage_df.iterrows():
        r += 1
        ws.cell(r, 1, row.get("stage_label", row["dealstage"]))
        ws.cell(r, 2, int(row["total_deals"]))
        ws.cell(r, 3, int(row["active_this_week"]))
    xl_style_data(ws, ds, r, 3)
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_cambios(wb, week_label, trans_df):
    ws = wb.create_sheet("🔄 Cambios de etapa")
    xl_header(ws, "Cambios de etapa detectados", f"Semana: {week_label}", 7)
    r = 4
    if trans_df.empty:
        ws.cell(r, 1, "No se detectaron cambios de etapa en este período.").font = NORMAL_FONT
        return
    for ci, h in enumerate(["Negocio","ID","Comercial","Pipeline","Etapa anterior","Etapa nueva","Fecha"], 1):
        ws.cell(r, ci, h)
    xl_style_headers(ws, r, 7)
    ds = r+1
    for _, row in trans_df.iterrows():
        r += 1
        ws.cell(r, 1, row.get("dealname",""))
        ws.cell(r, 2, row.get("deal_id",""))
        ws.cell(r, 3, row.get("owner",""))
        ws.cell(r, 4, row.get("pipeline",""))
        ws.cell(r, 5, row.get("from_stage",""))
        ws.cell(r, 6, row.get("to_stage",""))
        ts = row.get("changed_at")
        ws.cell(r, 7, str(ts)[:16] if ts else "")
    xl_style_data(ws, ds, r, 7)
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_nuevos(wb, week_label, new_df):
    ws = wb.create_sheet("✨ Nuevos negocios")
    xl_header(ws, "Negocios creados esta semana", f"Semana: {week_label}", 2)
    r = 4
    if new_df.empty:
        ws.cell(r, 1, "No se crearon negocios nuevos en este período.").font = NORMAL_FONT
        return
    for ci, h in enumerate(["Etapa inicial","Negocios nuevos"], 1):
        ws.cell(r, ci, h)
    xl_style_headers(ws, r, 2)
    ds = r+1
    for _, row in new_df.iterrows():
        r += 1
        ws.cell(r, 1, row.get("stage_label", row["dealstage"]))
        ws.cell(r, 2, int(row["new_deals"]))
    xl_style_data(ws, ds, r, 2)
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_estancados(wb, stale_data):
    ws = wb.create_sheet("⚠️ Estancados")
    xl_header(ws, "Negocios sin actividad", "Clasificados por días de inactividad", 5)
    r = 4
    for days, df_stale in stale_data.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1, f"SIN ACTIVIDAD MÁS DE {days} DÍAS ({len(df_stale)} negocios)").fill = SUBHEAD_FILL
        ws.cell(r, 1).font = SUBHEAD_FONT
        r += 1
        if df_stale.empty:
            ws.cell(r, 1, "Ninguno").font = NORMAL_FONT
            r += 2
            continue
        for ci, h in enumerate(["Negocio","Comercial","Etapa actual","Pipeline","Días sin actividad"], 1):
            ws.cell(r, ci, h)
        xl_style_headers(ws, r, 5)
        ds = r+1
        for _, row in df_stale.head(50).iterrows():
            r += 1
            ws.cell(r, 1, row.get("dealname",""))
            ws.cell(r, 2, row.get("owner",""))
            ws.cell(r, 3, row.get("stage_label",""))
            ws.cell(r, 4, row.get("pipeline",""))
            d = int(row.get("days_inactive", 0))
            ws.cell(r, 5, d)
            ws.cell(r, 5).fill = RED_FILL if d >= 60 else (ORANGE_FILL if d >= 30 else PatternFill())
        xl_style_data(ws, ds, r, 5)
        r += 2
    xl_auto_width(ws)
    ws.freeze_panes = "A5"

def xl_sheet_matriz(wb, matrix_df):
    ws = wb.create_sheet("🗃️ Matriz comercial-etapa")
    xl_header(ws, "Distribución total por comercial y etapa", "Todos los negocios", len(matrix_df.columns))
    r = 4
    for ci, col in enumerate(matrix_df.columns, 1):
        ws.cell(r, ci, col)
    xl_style_headers(ws, r, len(matrix_df.columns))
    for _, row in matrix_df.iterrows():
        r += 1
        for ci, val in enumerate(row, 1):
            cell = ws.cell(r, ci, val)
            cell.font      = BOLD_FONT if (ci == 1 or ci == len(row)) else NORMAL_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == len(row):
                cell.fill = ACCENT_FILL
    xl_auto_width(ws, min_w=8, max_w=30)
    ws.freeze_panes = "B5"

def generate_excel(week_label, act_df, stage_df, trans_df, new_df, matrix_df, stale_data):
    wb = Workbook()
    wb.remove(wb.active)
    xl_sheet_resumen(wb, week_label, act_df, stage_df, trans_df, new_df, stale_data)
    xl_sheet_actividad(wb, week_label, act_df)
    xl_sheet_etapas(wb, week_label, stage_df)
    xl_sheet_cambios(wb, week_label, trans_df)
    xl_sheet_nuevos(wb, week_label, new_df)
    xl_sheet_estancados(wb, stale_data)
    xl_sheet_matriz(wb, matrix_df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────
# HELPERS DE VISUALIZACIÓN
# ─────────────────────────────────────────────

def show_deal_table(df, title="Negocios", eng_types=None):
    """Muestra tabla de negocios con filtros por columna y buscador."""
    if df.empty:
        st.info("No hay negocios que mostrar.")
        return
    df2 = df.copy()
    df2["stage_label"]      = df2["dealstage"].map(STAGE_LABELS).fillna(df2["dealstage"])
    df2["Última actividad"] = df2["notes_last_updated"].dt.strftime("%d/%m/%Y %H:%M").fillna("—")

    ACT_EMOJI = {
        "Reunión": "📅 Reunión",
        "Email":   "📧 Email",
        "Llamada": "📞 Llamada",
        "Nota":    "📝 Nota",
        "—":       "—",
    }
    if eng_types:
        df2["Tipo actividad"] = df2["deal_id"].apply(
            lambda x: ACT_EMOJI.get(eng_types.get(str(x), "—"), "—")
        )
        cols   = ["dealname","owner","stage_label","pipeline","Tipo actividad","Última actividad"]
        rename = {"dealname":"Negocio","owner":"Comercial","stage_label":"Etapa","pipeline":"Pipeline"}
    else:
        cols   = ["dealname","owner","stage_label","pipeline","Última actividad"]
        rename = {"dealname":"Negocio","owner":"Comercial","stage_label":"Etapa","pipeline":"Pipeline"}

    show = df2[cols].rename(columns=rename)

    # ── Filtros ──────────────────────────────────
    uid = str(abs(hash(title)))[:6]
    with st.expander("🔍 Filtros", expanded=False):
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            buscar = st.text_input("🔎 Buscar negocio", key=f"buscar_{uid}", placeholder="Escribe para buscar...")
        with fc2:
            owners_opts = ["Todos"] + sorted(show["Comercial"].dropna().unique().tolist())
            owner_sel   = st.selectbox("👤 Comercial", options=owners_opts, key=f"owner_{uid}")
        with fc3:
            etapas_opts = ["Todas"] + sorted(show["Etapa"].dropna().unique().tolist())
            etapa_sel   = st.selectbox("📋 Etapa", options=etapas_opts, key=f"etapa_{uid}")

        fc4, fc5 = st.columns(2)
        with fc4:
            pipe_opts = ["Todos"] + sorted(show["Pipeline"].dropna().unique().tolist())
            pipe_sel  = st.selectbox("🏗️ Pipeline", options=pipe_opts, key=f"pipe_{uid}")
        with fc5:
            if "Tipo actividad" in show.columns:
                tipo_opts = ["Todos"] + sorted(show["Tipo actividad"].dropna().unique().tolist())
                tipo_sel  = st.selectbox("⚡ Tipo actividad", options=tipo_opts, key=f"tipo_{uid}")
            else:
                tipo_sel = "Todos"

    # Aplicar filtros
    filtered = show.copy()
    if buscar:
        filtered = filtered[filtered["Negocio"].str.contains(buscar, case=False, na=False)]
    if owner_sel != "Todos":
        filtered = filtered[filtered["Comercial"] == owner_sel]
    if etapa_sel != "Todas":
        filtered = filtered[filtered["Etapa"] == etapa_sel]
    if pipe_sel != "Todos":
        filtered = filtered[filtered["Pipeline"] == pipe_sel]
    if tipo_sel != "Todos" and "Tipo actividad" in filtered.columns:
        filtered = filtered[filtered["Tipo actividad"] == tipo_sel]

    st.caption(f"**{len(filtered)}** negocios" + (f" (de {len(show)} totales)" if len(filtered) != len(show) else ""))
    st.dataframe(
        filtered,
        hide_index=True,
        use_container_width=True,
        height=min(450, 55 + len(filtered) * 35),
        column_config={
            "Negocio":         st.column_config.TextColumn("Negocio", width="large"),
            "Comercial":       st.column_config.TextColumn("Comercial"),
            "Etapa":           st.column_config.TextColumn("Etapa", width="large"),
            "Pipeline":        st.column_config.TextColumn("Pipeline", width="small"),
            "Última actividad":st.column_config.TextColumn("Última actividad"),
        }
    )

def semaforo_color(rate):
    if rate >= 20:   return "#1D9E75"
    elif rate >= 10: return "#F4A835"
    else:            return "#E05252"

def semaforo_emoji(rate):
    if rate >= 20:   return "🟢"
    elif rate >= 10: return "🟡"
    else:            return "🔴"

PLOT_LAYOUT = dict(
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Arial", size=12),
)

# ─────────────────────────────────────────────
# APP STREAMLIT
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Pipeline Tracker — HubSpot",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stMetricValue"] { font-size: 1.8rem; font-weight: 700; }
[data-testid="stMetricLabel"] { font-size: 0.85rem; color: #666; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
div[data-testid="metric-container"] {
    background: #f8f9fb;
    border: 1px solid #e8eaf0;
    border-radius: 10px;
    padding: 16px;
}
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=3600, show_spinner=False)
def cached_all_deals(api_key):
    return parse_deals(get_all_deals(api_key))

@st.cache_data(ttl=3600, show_spinner=False)
def cached_stage_history(api_key, deal_ids_tuple):
    return get_stage_history_bulk(api_key, list(deal_ids_tuple))

@st.cache_data(ttl=3600, show_spinner=False)
def cached_real_activity(api_key, since_ts_ms, until_ts_ms):
    return get_real_activity(api_key, since_ts_ms, until_ts_ms)

@st.cache_data(ttl=3600, show_spinner=False)
def cached_engagement_types(api_key, deal_ids_tuple):
    return get_engagement_types(api_key, list(deal_ids_tuple))

# ── Barra lateral ──────────────────────────────

api_key = HUBSPOT_API_KEY

st.sidebar.title("📊 Pipeline Tracker")
st.sidebar.caption("Seguimiento semanal de HubSpot")
st.sidebar.divider()

st.sidebar.subheader("Período de análisis")
col1, col2 = st.sidebar.columns(2)
with col1:
    start_date = st.date_input("Desde", value=datetime.now().date() - timedelta(days=7))
with col2:
    end_date = st.date_input("Hasta", value=datetime.now().date())

st.sidebar.divider()

load_history = st.sidebar.checkbox(
    "Cargar historial de etapas", value=True,
    help="Necesario para ver cambios de etapa.",
)
load_real_activity = st.sidebar.checkbox(
    "Desglose por tipo de actividad", value=False,
    help="Carga llamadas, emails, reuniones, notas y tareas. Más lento.",
)
pipeline_filter = st.sidebar.multiselect(
    "Filtrar pipeline", options=list(PIPELINE_NAMES.values()),
    default=list(PIPELINE_NAMES.values()),
)
owner_filter = st.sidebar.multiselect(
    "Filtrar comercial", options=sorted(OWNER_NAMES.values()),
    default=[], placeholder="Todos",
)

st.sidebar.divider()
st.sidebar.caption("🟢 Actividad ≥20%  🟡 10-20%  🔴 <10%")
st.sidebar.caption("Actividad = notes_last_updated. Tipos: reuniones, emails, llamadas y notas (tareas excluidas).")

# ── Login ──────────────────────────────────────

if APP_PASSWORD:
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.markdown("<br>" * 3, unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 1.5, 1])
        with col2:
            st.markdown("""
            <div style="text-align:center; margin-bottom:24px;">
                <div style="font-size:2.5rem">📊</div>
                <div style="font-size:1.4rem; font-weight:700; color:#1A3A5C">Pipeline Tracker</div>
                <div style="font-size:13px; color:#888; margin-top:4px">Zentralcom · HubSpot</div>
            </div>
            """, unsafe_allow_html=True)
            pwd = st.text_input("Contraseña", type="password", placeholder="Introduce la contraseña...")
            if st.button("Entrar", use_container_width=True):
                if pwd == APP_PASSWORD:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
        st.stop()

# ── Comprobación API ───────────────────────────

if not api_key or api_key.startswith("pat-eu1-xxxxxxxx"):
    st.error("⚠️ Pon tu API Key real en la variable HUBSPOT_API_KEY al inicio del código.")
    st.stop()

if not test_connection(api_key):
    st.error("❌ No se pudo conectar con HubSpot. Revisa tu API Key.")
    st.stop()

# ── Carga de datos ─────────────────────────────

week_start = datetime.combine(start_date, datetime.min.time())
week_end   = datetime.combine(end_date,   datetime.max.time())
week_label = f"{start_date.strftime('%d/%m/%Y')} — {end_date.strftime('%d/%m/%Y')}"
today      = datetime.now()

with st.spinner("Cargando negocios de HubSpot..."):
    df_all = cached_all_deals(api_key)

if pipeline_filter:
    df_all = df_all[df_all["pipeline"].isin(pipeline_filter)]
if owner_filter:
    df_all = df_all[df_all["owner"].isin(owner_filter)]

# A partir de aquí todo el dashboard trabaja SOLO con negocios abiertos
df_closed = df_all[df_all["dealstage"].isin(CLOSED_STAGES)].copy()  # guardamos cerrados solo para cierres KPI
df_all    = df_all[~df_all["dealstage"].isin(CLOSED_STAGES)].copy()

# df_week basado en notes_last_updated — misma propiedad que usa HubSpot UI
df_week = df_all[
    df_all["notes_last_updated"].notna() &
    (df_all["notes_last_updated"] >= week_start) &
    (df_all["notes_last_updated"] <= week_end)
].copy()

trans_df = pd.DataFrame(columns=["deal_id","dealname","owner","pipeline","from_stage","to_stage","changed_at"])
if load_history and not df_week.empty:
    with st.spinner(f"Cargando historial de etapas para {len(df_week)} negocios..."):
        history_data = cached_stage_history(api_key, tuple(df_week["deal_id"].tolist()))
    trans_df = stage_transitions(history_data, week_start, week_end)

activity_detail_df = pd.DataFrame(columns=["owner","tipo"])
engagement_types   = {}
if not df_week.empty:
    with st.spinner(f"Cargando tipo de actividad para {len(df_week)} negocios..."):
        engagement_types = cached_engagement_types(api_key, tuple(df_week["deal_id"].tolist()))
    rows = []
    for _, row in df_week.iterrows():
        tipo = engagement_types.get(str(row["deal_id"]), "—")
        if tipo != "—":
            rows.append({"owner": row["owner"], "tipo": tipo})
    activity_detail_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["owner","tipo"])

act_df     = activity_by_owner(df_all, week_start, week_end)
stage_df   = stage_activity_summary(df_all, week_start, week_end)
new_df     = new_deals_by_stage(df_all, week_start, week_end)
matrix_df  = owner_stage_matrix(df_all)
stale_data = stale_deals(df_all, STALE_DAYS, today)

# ── CABECERA ───────────────────────────────────

st.title("📊 Pipeline Tracker — HubSpot")
st.caption(f"Período: **{week_label}** · {len(df_all):,} negocios abiertos · Actualizado: {today.strftime('%d/%m/%Y %H:%M')}")

# ── KPIs ───────────────────────────────────────

total      = len(df_all)  # df_all ya es solo abiertos
active     = int(act_df["active_this_week"].sum())
rate       = round(active/total*100, 1) if total else 0
# Cierres = negocios cerrados con fecha de cierre en el período (usando df_closed)
won  = len(df_closed[
    df_closed["dealstage"].isin(WON_STAGES) &
    df_closed["closedate"].notna() &
    (df_closed["closedate"] >= week_start) &
    (df_closed["closedate"] <= week_end)
])
lost = len(df_closed[
    df_closed["dealstage"].isin(LOST_STAGES) &
    df_closed["closedate"].notna() &
    (df_closed["closedate"] >= week_start) &
    (df_closed["closedate"] <= week_end)
])
new_total = int(new_df["new_deals"].sum()) if not new_df.empty else 0
stale14   = len(stale_data.get(14, pd.DataFrame()))
stale30   = len(stale_data.get(30, pd.DataFrame()))

c = st.columns(7)
# KPIs clickables
kpi_data = [
    ("Negocios abiertos",   f"{total:,}",          None,                     "total"),
    ("Con actividad",       f"{active:,}",          f"{rate}% del total",     "activos"),
    ("Tasa actividad",      f"{rate}%",             None,                     "activos"),
    ("Cambios de etapa",    f"{len(trans_df):,}",   None,                     "cambios"),
    ("Nuevos en período",   f"{new_total:,}",        None,                    "nuevos"),
    ("Cierres ganados",     f"{won:,}",             f"{lost} perdidos en período","ganados"),
    ("Estancados >14d",     f"{stale14:,}",         f"De {total} abiertos · {stale30} >30d", "estancados"),
]
for col, (label, val, delta, key) in zip(c, kpi_data):
    col.metric(label, val, delta)

# Desplegables KPIs
nuevos_df = df_all[(df_all["createdate"] >= week_start) & (df_all["createdate"] <= week_end)]
ganados_df = df_closed[
    df_closed["dealstage"].isin(WON_STAGES) &
    df_closed["closedate"].notna() &
    (df_closed["closedate"] >= week_start) &
    (df_closed["closedate"] <= week_end)
]
perdidos_df = df_closed[
    df_closed["dealstage"].isin(LOST_STAGES) &
    df_closed["closedate"].notna() &
    (df_closed["closedate"] >= week_start) &
    (df_closed["closedate"] <= week_end)
]

with st.expander(f"🔵 **Negocios abiertos** — {total:,} en el pipeline"):
    show_deal_table(df_all)

with st.expander(f"✅ **Con actividad** — {active:,} en el período seleccionado"):
    show_deal_table(df_week, eng_types=engagement_types)

with st.expander(f"✨ **Nuevos en período** — {new_total:,} creados"):
    show_deal_table(nuevos_df)

with st.expander(f"🏆 **Cierres** — {won:,} ganados / {lost:,} perdidos en el período"):
    if not ganados_df.empty:
        st.markdown("**✅ Ganados**")
        show_deal_table(ganados_df, eng_types=engagement_types)
    if not perdidos_df.empty:
        st.markdown("**❌ Perdidos**")
        show_deal_table(perdidos_df, eng_types=engagement_types)
    if ganados_df.empty and perdidos_df.empty:
        st.info("No hay cierres en el período seleccionado.")

with st.expander(f"⚠️ **Estancados >14 días** — {stale14:,} negocios"):
    show_deal_table(stale_data.get(14, pd.DataFrame()))

st.divider()

# ── PESTAÑAS ───────────────────────────────────

tabs = st.tabs([
    "👤 Comerciales",
    "📋 Etapas",
    "🔄 Cambios de etapa",
    "✨ Nuevos negocios",
    "⚠️ Estancados",
    "🌡️ Mapa de calor",
])

# ── TAB 1: COMERCIALES ─────────────────────────

with tabs[0]:
    st.subheader("Actividad semanal por comercial")

    # Tarjetas semáforo
    cols = st.columns(max(len(act_df), 1))
    for col, (_, row) in zip(cols, act_df.iterrows()):
        color = semaforo_color(row["activity_rate"])
        emoji = semaforo_emoji(row["activity_rate"])
        pct   = row["activity_rate"]
        bg    = "#f0faf6" if pct >= 20 else "#fffbf0" if pct >= 10 else "#fdf2f2"
        with col:
            st.markdown(f"""
            <div style="border:1.5px solid {color}; border-radius:12px; padding:14px 10px;
                        text-align:center; background:{bg}">
                <div style="font-size:1.5rem">{emoji}</div>
                <div style="font-weight:700; font-size:0.9rem; margin:4px 0">{row['owner'].split()[0]}</div>
                <div style="font-size:1.6rem; font-weight:800; color:{color}">{pct}%</div>
                <div style="font-size:0.78rem; color:#888">{row['active_this_week']} / {row['open_deals']} abiertos</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Gráfico barras apiladas
    fig = go.Figure()
    fig.add_bar(
        name="Sin actividad (abiertos)",
        x=act_df["owner"], y=act_df["sin_actividad"],
        marker_color="#E8EAF0", marker_line_width=0,
    )
    fig.add_bar(
        name="Con actividad esta semana",
        x=act_df["owner"], y=act_df["active_this_week"],
        marker_color=[semaforo_color(r) for r in act_df["activity_rate"]],
        marker_line_width=0,
        text=act_df["active_this_week"], textposition="outside",
    )
    fig.update_layout(
        barmode="stack", height=380,
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        **PLOT_LAYOUT,
    )
    fig.update_xaxes(tickangle=-20)
    st.plotly_chart(fig, use_container_width=True)

    # Barras de progreso con expander clickable
    st.markdown("##### Detalle — haz clic en un comercial para ver sus negocios")
    for _, row in act_df.iterrows():
        pct   = row["activity_rate"]
        color = semaforo_color(pct)
        bar_w = min(int(pct * 3), 300)
        emoji = semaforo_emoji(row["activity_rate"])

        with st.expander(
            f"{emoji} **{row['owner']}** — {row['active_this_week']} activos / {row['total_deals']} total ({pct}%)",
            expanded=False
        ):
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown(f"""
                <div style="background:#f8f9fb; border-radius:8px; padding:12px 16px;">
                    <div style="display:flex; align-items:center; gap:10px;">
                        <div style="flex:1; background:#E8EAF0; border-radius:4px; height:14px;">
                            <div style="width:{bar_w}px; max-width:100%; background:{color};
                                        height:14px; border-radius:4px;"></div>
                        </div>
                        <div style="font-weight:700; color:{color}">{pct}%</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            tab_a, tab_b = st.tabs(["✅ Con actividad en el período", "💤 Sin actividad en el período"])
            with tab_a:
                deals_active = df_week[df_week["owner"] == row["owner"]]
                show_deal_table(deals_active, f"Con actividad — {row['owner']}", eng_types=engagement_types)
            with tab_b:
                deals_inactive = df_all[
                    (df_all["owner"] == row["owner"]) &
                    ~(df_all["deal_id"].isin(df_week["deal_id"]))
                ]
                show_deal_table(deals_inactive, f"Sin actividad — {row['owner']}")

    # Desglose por tipo (solo si se activó)
    if load_real_activity:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("##### Desglose por tipo de actividad")
        if activity_detail_df.empty:
            st.info("No se encontró actividad real asociada a negocios en este período.")
        else:
            TYPE_LABELS = {
                "calls":    "📞 Llamadas",
                "emails":   "📧 Emails",
                "meetings": "📅 Reuniones",
                "notes":    "📝 Notas",
                "tasks":    "✅ Tareas",
            }
            if "activity_type" in activity_detail_df.columns:
                activity_detail_df["tipo"] = activity_detail_df["activity_type"].map(TYPE_LABELS).fillna(activity_detail_df["activity_type"])
            pivot_act = activity_detail_df.groupby(["owner","tipo"]).size().reset_index(name="n")
            fig2 = px.bar(
                pivot_act, x="owner", y="n", color="tipo",
                barmode="stack",
                labels={"owner":"Comercial","n":"Actividades","tipo":"Tipo"},
                color_discrete_map={
                    "📞 Llamadas":  "#185FA5",
                    "📧 Emails":    "#1D9E75",
                    "📅 Reuniones": "#F4A835",
                    "📝 Notas":     "#9B59B6",
                    "✅ Tareas":    "#E05252",
                },
                height=340,
            )
            fig2.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02), **PLOT_LAYOUT)
            fig2.update_xaxes(tickangle=-20)
            st.plotly_chart(fig2, use_container_width=True)

            type_totals = activity_detail_df["tipo"].value_counts().reset_index()
            type_totals.columns = ["Tipo","Total"]
            c_cols = st.columns(len(type_totals))
            for col, (_, row) in zip(c_cols, type_totals.iterrows()):
                col.metric(row["Tipo"], int(row["Total"]))

# ── TAB 2: ETAPAS ──────────────────────────────

with tabs[1]:
    st.subheader("Actividad por etapa esta semana")
    top = stage_df[stage_df["active_this_week"] > 0].head(20)
    if top.empty:
        st.info("No hay actividad en ninguna etapa en el período seleccionado.")
    else:
        fig = px.bar(
            top, x="active_this_week", y="stage_label", orientation="h",
            color="active_this_week",
            color_continuous_scale=["#B5D4F4","#185FA5"],
            text="active_this_week",
            labels={"active_this_week":"Negocios con actividad","stage_label":""},
            height=520,
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(coloraxis_showscale=False, margin=dict(t=20,b=20,l=260,r=60), **PLOT_LAYOUT)
        st.plotly_chart(fig, use_container_width=True)

    with st.expander("Ver tabla completa de etapas"):
        st.dataframe(
            stage_df[["stage_label","total_deals","active_this_week"]].rename(columns={
                "stage_label":"Etapa","total_deals":"Total negocios","active_this_week":"Con actividad"
            }),
            hide_index=True, use_container_width=True,
        )

    st.markdown("##### Haz clic en una etapa para ver sus negocios")
    for _, srow in stage_df[stage_df["active_this_week"] > 0].head(20).iterrows():
        label = srow.get("stage_label", srow["dealstage"])
        with st.expander(f"📋 **{label}** — {srow['active_this_week']} con actividad / {srow['total_deals']} total"):
            tab_s1, tab_s2 = st.tabs(["✅ Con actividad en el período", "💤 Todos en esta etapa"])
            with tab_s1:
                deals_stage_active = df_week[df_week["dealstage"] == srow["dealstage"]]
                show_deal_table(deals_stage_active, eng_types=engagement_types)
            with tab_s2:
                deals_stage_all = df_all[df_all["dealstage"] == srow["dealstage"]]
                show_deal_table(deals_stage_all)

# ── TAB 3: CAMBIOS DE ETAPA ────────────────────

with tabs[2]:
    st.subheader("Cambios de etapa esta semana")
    if trans_df.empty:
        st.info("No se detectaron cambios de etapa. Activa 'Cargar historial de etapas' en la barra lateral.")
    else:
        st.caption(f"**{len(trans_df)}** cambios detectados en el período")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("##### 🟢 Etapas con más entradas")
            to_c = trans_df.groupby("to_stage").size().reset_index(name="n").sort_values("n", ascending=False).head(10)
            fig = px.bar(to_c, x="n", y="to_stage", orientation="h", text="n",
                         color_discrete_sequence=["#1D9E75"],
                         labels={"n":"","to_stage":""}, height=320)
            fig.update_traces(textposition="outside")
            fig.update_layout(margin=dict(t=10,b=10,l=230,r=50), **PLOT_LAYOUT)
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            st.markdown("##### 🔴 Etapas con más salidas")
            from_c = trans_df.groupby("from_stage").size().reset_index(name="n").sort_values("n", ascending=False).head(10)
            fig = px.bar(from_c, x="n", y="from_stage", orientation="h", text="n",
                         color_discrete_sequence=["#E05252"],
                         labels={"n":"","from_stage":""}, height=320)
            fig.update_traces(textposition="outside")
            fig.update_layout(margin=dict(t=10,b=10,l=230,r=50), **PLOT_LAYOUT)
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("##### Detalle de todos los cambios")

        # Agrupar movimientos por negocio en una sola línea
        trans_sorted = trans_df.sort_values(["dealname","changed_at"]).reset_index(drop=True)

        # Construir recorrido completo por negocio
        deals_flow = {}
        for _, row in trans_sorted.iterrows():
            did = row["deal_id"]
            if did not in deals_flow:
                deals_flow[did] = {
                    "dealname": row["dealname"],
                    "owner":    row["owner"],
                    "pipeline": row["pipeline"],
                    "fecha":    str(row["changed_at"])[:16] if row["changed_at"] else "—",
                    "stages":   [row["from_stage"]],
                }
            deals_flow[did]["stages"].append(row["to_stage"])
            # Actualizar fecha con la más reciente
            if row["changed_at"]:
                deals_flow[did]["fecha"] = str(row["changed_at"])[:16]

        for did, info in deals_flow.items():
            stages    = info["stages"]
            n_moves   = len(stages) - 1
            badge     = f'<span style="background:#E8EAF0;border-radius:10px;padding:2px 8px;font-size:11px;color:#555;margin-left:8px">{n_moves} movimiento{"s" if n_moves>1 else ""}</span>'

            # Construir la cadena de etapas con flechas
            stages_html = ""
            for i, stage in enumerate(stages):
                if i == 0:
                    # Primera etapa: naranja (origen)
                    color_bg, color_txt = "#FCE4D6", "#7B3B00"
                elif i == len(stages) - 1:
                    # Última etapa: verde (destino final)
                    color_bg, color_txt = "#E2EFDA", "#375623"
                else:
                    # Etapas intermedias: azul claro
                    color_bg, color_txt = "#E6F1FB", "#0C447C"
                stages_html += f'<span style="background:{color_bg};color:{color_txt};padding:3px 10px;border-radius:12px;font-size:12px;white-space:nowrap">{stage}</span>'
                if i < len(stages) - 1:
                    stages_html += '<span style="font-size:16px;color:#185FA5;margin:0 4px">→</span>'

            st.markdown(f"""
            <div style="margin-top:10px; padding:10px 14px; background:#f8f9fb;
                        border-radius:8px; border-left:3px solid #185FA5;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                    <span style="font-weight:700; font-size:0.9rem">
                        {info['dealname']}{badge}
                    </span>
                    <span style="font-size:11px; color:#888">{info['owner']} · {info['pipeline']} · {info['fecha']}</span>
                </div>
                <div style="display:flex; align-items:center; gap:4px; flex-wrap:wrap;">
                    {stages_html}
                </div>
            </div>
            """, unsafe_allow_html=True)

# ── TAB 4: NUEVOS NEGOCIOS ─────────────────────

with tabs[3]:
    st.subheader("Negocios creados esta semana")
    if new_df.empty:
        st.info("No se crearon negocios nuevos en el período seleccionado.")
    else:
        c1, c2 = st.columns([1, 2])
        with c1:
            st.metric("Total nuevos esta semana", int(new_df["new_deals"].sum()))
            st.dataframe(
                new_df[["stage_label","new_deals"]].rename(columns={
                    "stage_label":"Etapa inicial","new_deals":"Negocios"
                }),
                hide_index=True, use_container_width=True,
            )
        with c2:
            fig = px.pie(
                new_df, values="new_deals", names="stage_label", hole=0.45,
                color_discrete_sequence=px.colors.sequential.Blues_r,
            )
            fig.update_traces(textposition="inside", textinfo="percent+label")
            fig.update_layout(showlegend=False, height=380, **PLOT_LAYOUT)
            st.plotly_chart(fig, use_container_width=True)

# ── TAB 5: ESTANCADOS ──────────────────────────

with tabs[4]:
    st.subheader("⚠️ Negocios sin actividad")
    st.caption("Negocios abiertos (no ganados ni perdidos) sin actividad real desde hace X días.")

    threshold = st.radio(
        "Mostrar negocios sin actividad en más de:",
        options=[14, 30, 60],
        format_func=lambda x: f"{x} días",
        horizontal=True,
    )
    df_stale = stale_data.get(threshold, pd.DataFrame())

    if df_stale.empty:
        st.success(f"✅ No hay negocios estancados más de {threshold} días.")
    else:
        st.warning(f"**{len(df_stale)} negocios** llevan más de {threshold} días sin actividad.")
        by_owner = df_stale.groupby("owner").size().reset_index(name="estancados").sort_values("estancados", ascending=False)
        fig = px.bar(
            by_owner, x="owner", y="estancados", text="estancados",
            color="estancados", color_continuous_scale=["#FFF2CC","#E05252"],
            labels={"owner":"Comercial","estancados":"Negocios estancados"},
            height=320,
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(coloraxis_showscale=False, **PLOT_LAYOUT)
        st.plotly_chart(fig, use_container_width=True)

        def color_dias(val):
            if val >= 60:  return "background-color: #fde8e8; color: #a00"
            elif val >= 30: return "background-color: #fff3cd; color: #7a5000"
            return ""

        # Tabla global con colores
        show_df = df_stale[["dealname","owner","stage_label","pipeline","days_inactive"]].rename(columns={
            "dealname":"Negocio","owner":"Comercial","stage_label":"Etapa actual",
            "pipeline":"Pipeline","days_inactive":"Días sin actividad",
        })
        try:
            styled = show_df.style.map(color_dias, subset=["Días sin actividad"])
        except AttributeError:
            styled = show_df.style.applymap(color_dias, subset=["Días sin actividad"])
        st.dataframe(
            styled,
            hide_index=True, use_container_width=True, height=350,
        )

        # Detalle por comercial
        st.markdown("##### Detalle por comercial")
        for owner_name, grp in df_stale.groupby("owner"):
            with st.expander(f"👤 **{owner_name}** — {len(grp)} negocios estancados"):
                show_deal_table(grp)

# ── TAB 6: MAPA DE CALOR ───────────────────────

with tabs[5]:
    st.subheader("🌡️ Mapa de calor — Comercial × Etapa")
    st.caption("Intensidad = número de negocios. Detecta dónde se acumulan y posibles cuellos de botella.")

    df_heat = df_all.copy()
    df_heat["stage_label"] = df_heat["dealstage"].map(STAGE_LABELS).fillna(df_heat["dealstage"])
    stage_counts = df_heat.groupby("stage_label")["deal_id"].count()
    top_stages   = stage_counts[stage_counts > 0].nlargest(15).index.tolist()
    df_heat      = df_heat[df_heat["stage_label"].isin(top_stages)]
    pivot = df_heat.pivot_table(index="owner", columns="stage_label",
                                values="deal_id", aggfunc="count", fill_value=0)
    # Añadir fila de totales por etapa
    pivot_total = pivot.copy()
    totals = pivot_total.sum(axis=0)
    totals.name = "TOTAL"
    pivot_with_total = pd.concat([pivot_total, totals.to_frame().T])

    fig = px.imshow(
        pivot_with_total,
        color_continuous_scale=["#EBF4FB","#185FA5"],
        aspect="auto", text_auto=True,
        labels={"color":"Negocios"},
        height=460,
    )
    fig.update_xaxes(tickangle=-35, tickfont_size=11)
    fig.update_yaxes(tickfont_size=11)
    fig.update_layout(
        coloraxis_showscale=True,
        margin=dict(t=30, b=120, l=150, r=20),
        **PLOT_LAYOUT,
    )
    # Destacar la fila TOTAL con borde
    fig.add_hline(
        y=len(pivot_with_total) - 1.5,
        line_dash="dash", line_color="#185FA5", line_width=1.5,
    )
    st.plotly_chart(fig, use_container_width=True)

    # Tabla de totales por etapa debajo
    st.markdown("##### Total negocios por etapa")
    totals_df = pd.DataFrame({
        "Etapa": totals.index,
        "Total negocios": totals.values.astype(int)
    }).sort_values("Total negocios", ascending=False)
    st.dataframe(totals_df, hide_index=True, use_container_width=True, height=300)



# ── EXPORTAR EXCEL ─────────────────────────────

st.divider()
st.subheader("📥 Exportar a Excel")
st.caption("Excel con 7 pestañas: resumen, actividad, etapas, cambios, nuevos, estancados y matriz completa.")

filename    = f"pipeline_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
excel_bytes = generate_excel(week_label, act_df, stage_df, trans_df, new_df, matrix_df, stale_data)

c1, c2 = st.columns(2)
with c1:
    st.download_button(
        label="⬇️ Descargar en el navegador",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with c2:
    if st.button("💾 Guardar en carpeta local", use_container_width=True):
        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            full_path = os.path.join(OUTPUT_DIR, filename)
            with open(full_path, "wb") as f:
                f.write(excel_bytes)
            st.success(f"✅ Guardado en:\n`{full_path}`")
        except Exception as e:
            st.error(f"❌ Error al guardar: {e}")
